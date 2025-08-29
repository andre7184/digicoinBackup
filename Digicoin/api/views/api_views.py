from openpyxl import load_workbook
from rest_framework import viewsets, status
from api.models import *
from api.serializers import *
from rest_framework.views import APIView
from django.contrib.auth import authenticate, login, logout
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
import yagmail
import os
from dotenv import load_dotenv

class User(APIView):
    
    def get(self, request, id=None):
        if id:
            usuario = get_object_or_404(CustomUser, pk=id)
            serializer = UserSerializer(usuario)
            return Response(serializer.data, status=status.HTTP_200_OK)

        nome = request.query_params.get("nome")
        
        if nome:
            usuario = CustomUser.objects.filter(first_name__icontains=nome)[:5]
        else:
            usuario = CustomUser.objects.all()[:5]

        serializer = UserSerializer(usuario, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



    def post(self, request):
        nome = request.data.get('nome')
        senha = request.data.get('senha')
        ra = request.data.get('ra')
        fistName = request.data.get('first_name')
        isAdm = request.data.get('is_adm')
        isActive = request.data.get('is_active')

        if not nome or not senha:
            return Response({"error": "Todos os campos são obrigatórios!", "status": status.HTTP_400_BAD_REQUEST}, status= status.HTTP_400_BAD_REQUEST)

        usuario = CustomUser.objects.create(
            username = nome,
            password = make_password(senha),
            first_name = fistName,
            is_adm = isAdm,
            is_active = isActive,
            ra = ra
        )

        # # Enviar email após cadastro #
        # load_dotenv()

        # yag = yagmail.SMTP(
        #     user=os.getenv("EMAIL_USER"),
        #     password=os.getenv("EMAIL_PASSWORD"),
        #     host=os.getenv("EMAIL_HOST"),
        #     port=int(os.getenv("EMAIL_PORT", "587")),
        #     smtp_starttls=True,       
        #     smtp_ssl=False    
        # )

        # yag.send(
        #     to=usuario.username,
        #     subject='Bem vindo ao Sistema Digicoin',
        #     contents=(
        #         f'Nome: {fistName}\n'
        #         f'RA: {ra}\n'
        #         f'Login: {nome}\n'
        #         f'Senha: {senha}\n'
        #         f"Altere sua senha depois do primeiro acesso."
        #     )
        # )

        return Response({"message":"Usuário criado com sucesso!", "id":usuario.id, "status": status.HTTP_201_CREATED})

    def put(self, request, id):
        usuario = get_object_or_404(CustomUser, pk=id)
        data = request.data.copy()
        operacao = data.get("operacao")
       
        
        if operacao in ['adicionar', 'remover']:
            try:
                saldo = int(data.get("saldo", 0))
            except (TypeError, ValueError):
                return Response({"erro": "Saldo inválido."}, status=status.HTTP_400_BAD_REQUEST)

            if operacao == 'adicionar':
                usuario.pontuacao += saldo
                usuario.saldo += saldo

            elif operacao == 'remover':
                if usuario.saldo < saldo:
                    return Response({"erro": "Saldo insuficiente."}, status=status.HTTP_400_BAD_REQUEST)
                usuario.pontuacao -= saldo
                usuario.saldo -= saldo

            usuario.save()
            return Response({"status": status.HTTP_200_OK})
        
        serializer = UserSerializer(usuario, data=data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({"status": status.HTTP_200_OK})

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        usuario = get_object_or_404(CustomUser, pk = id)
        if usuario:
            usuario.delete()
            return Response({"status": status.HTTP_200_OK})
        else:
            return Response({"status": status.HTTP_404_NOT_FOUND})
    
class PrimeiroAcessoSenhaView(APIView):
    def post(self, request, id):
        senha = request.data.get('senha')
        confirmar_senha = request.data.get('confirmarSenha')

        if not senha or not confirmar_senha:
            return Response({"erro": "Ambas as senhas são obrigatórias."}, status=status.HTTP_400_BAD_REQUEST)

        if senha != confirmar_senha:
            return Response({"erro": "As senhas não coincidem."}, status=status.HTTP_400_BAD_REQUEST)

        usuario = get_object_or_404(CustomUser, pk=id)
        usuario.password = make_password(senha)
        usuario.primeiroAcesso = False
        usuario.save()
        logout(request)
        return Response({"mensagem": "Senha atualizada com sucesso."}, status=status.HTTP_200_OK)


class Login(APIView):
    def post(self, request):
        nome = request.data.get('nome')
        senha = request.data.get('senha')
        
        user = authenticate(username=nome, password=senha)
        if user is not None:
            login(request, user)
            return Response({
                'is_adm': user.is_adm
            }, status=status.HTTP_200_OK)
        
        return Response({'error': 'Credenciais inválidas',}, status=status.HTTP_401_UNAUTHORIZED)
        
class Logout(APIView):
    def post(self, request):
        logout(request)
        return Response({"status": status.HTTP_200_OK, "mensagem": "Logout realizado com sucesso"})

class GetDadosUsuarioLogado(APIView):
    def get(self, request):
        usuarioId = request.session.get('_auth_user_id')
        if usuarioId:
            usuario = CustomUser.objects.filter(id= usuarioId).first()
            serializer = UserSerializer(usuario)
            return Response(serializer.data)

        return Response(usuarioId)
    

    
class CampanhaViewSet(viewsets.ModelViewSet):
    queryset = Campanha.objects.all()
    serializer_class = CampanhaSerializer

class ProdutoViewSet(viewsets.ModelViewSet):
    queryset = Produto.objects.all()
    serializer_class = ProdutoSerializer

class DesafioViewSet(viewsets.ModelViewSet):
    queryset = Desafio.objects.all()
    serializer_class = DesafioSerializer

class CompraViewSet(viewsets.ModelViewSet):

    queryset = Compra.objects.all()
    serializer_class = CompraSerializer

class ItensCompraViewSet(viewsets.ModelViewSet):
    queryset = ItensCompra.objects.all()
    serializer_class = ItensCompraSerializer

class CadastrarCompraView(APIView):
    def post(self, request):
        dadosCompra = request.data.get('compra')
        itensCompra = request.data.get('itens')
        usuario_id = request.user.id
        # adiciona o id do usuário ao dadosCompra
        dadosCompra['idUsuario'] = usuario_id

        # Verifica se o usuário tem moeda suficiente
        usuario = CustomUser.objects.get(id=usuario_id)
        total_custo = dadosCompra['total']
        if usuario.saldo < total_custo:
            return Response({"error": "Usuário não tem saldo suficiente.", "status":status.HTTP_400_BAD_REQUEST})

        # Verifica se os produtos têm quantidade suficiente
        erros = []
        for item in itensCompra:
            produto = Produto.objects.get(id=item['idProduto'])
            if produto.quantidade < item['qtdProduto']:
                erros.append(f"Produto {produto.nome} não tem quantidade suficiente.")
        
        if erros:
            return Response({"error": erros, "status":status.HTTP_400_BAD_REQUEST})

        # Cria a compra
        compraSerializer = CompraSerializer(data=dadosCompra)
        if compraSerializer.is_valid():
            compra = compraSerializer.save()
        else:
            return Response(compraSerializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Cria os itens de compra
        for item in itensCompra:
            item['idCompra'] = compra.id
            itemSerializer = ItensCompraSerializer(data=item)
            if itemSerializer.is_valid():
                itemSerializer.save()
                produto.quantidade -= item['qtdProduto']
                produto.save()
            else:
                erros.append(itemSerializer.errors)

        if erros:
            return Response(erros, status=status.HTTP_400_BAD_REQUEST)

        # Deduzir moeda do usuário
        usuario.saldo -= total_custo
        usuario.save()

        return Response({"message": "Compra e itens criados com sucesso!", "status": status.HTTP_201_CREATED})
    
 
class HistoricoSaldoUsuarioView(APIView):
    """Retorna as últimas 5 alterações de saldo do usuário logado"""
    def get(self, request):
        usuario = request.user
        serializer = UsuarioComHistoricoSerializer(usuario)
        return Response(serializer.data, status=status.HTTP_200_OK)

class HistoricoSaldoPorIdView(APIView):
    """Retorna as últimas 5 alterações de saldo de um usuário pelo ID"""
    def get(self, request, id):
        usuario = get_object_or_404(CustomUser, pk=id)
        serializer = UsuarioComHistoricoSerializer(usuario)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class AtualizarSaldos(APIView):
    def put(self, request):
        data = request.data
        operacao = data.get("operacao")
        saldo = data.get("saldo")
        para_todos = data.get("paraTodos", False)
        usuarios_ids = data.get("usuarios", [])

        if not operacao or saldo is None:
            return Response({"erro": "Dados incompletos.", "status": status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)

        if para_todos:
            usuarios = CustomUser.objects.filter(is_active=True, is_adm=False)
        else:
            usuarios = CustomUser.objects.filter(id__in=usuarios_ids, is_active=True, is_adm=False)

        for usuario in usuarios:
            if operacao == "adicionar":
                usuario.pontuacao += saldo
                usuario.saldo += saldo
            elif operacao == "remover":
                if usuario.saldo < saldo:
                    continue  # ou trate como erro
                usuario.pontuacao -= saldo
                usuario.saldo -= saldo
            usuario.save()
        
        return Response({"message": "Operação realizada com sucesso!", "status": status.HTTP_200_OK}, status=status.HTTP_200_OK)

class ValidarImportacaoUsuarios(APIView):
    def post(self, request):
        arquivo = request.FILES.get('arquivo_usuarios')
        if not arquivo:
            return Response({'erro': 'Nenhum arquivo enviado.', 'status': status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)

        usuarios_validados = []
        emails_existentes = set(CustomUser.objects.values_list('username', flat=True))

        try:
            if arquivo.name.endswith('.csv'):
                import csv
                from io import TextIOWrapper
                decoded_file = TextIOWrapper(arquivo.file, encoding='utf-8')
                reader = csv.DictReader(decoded_file)
                rows = list(reader)
            elif arquivo.name.endswith(('.xls', '.xlsx')):
                wb = load_workbook(filename=arquivo, data_only=True)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                rows = [
                    {headers[i]: cell.value for i, cell in enumerate(row)}
                    for row in sheet.iter_rows(min_row=2)
                ]
            else:
                return Response({'erro': 'Formato de arquivo inválido. Use CSV ou XLSX.', 'status': status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'erro': f'Erro ao ler o arquivo: {str(e)}', 'status': status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)

        for row in rows:
            email = row.get('email')
            if row.get('nome') and email and row.get('senha'):
                if email in emails_existentes:
                    continue
                usuarios_validados.append({
                    'first_name': row.get('nome'),
                    'username': email,
                    'senha': row.get('senha'),
                    'ra': row.get('ra', ''),
                    'saldo': int(row.get('saldo') or 0),
                    'pontuacao': int(row.get('pontuacao') or 0),
                    'is_adm': str(row.get('is_adm')).lower() == 'true',
                    'is_active': str(row.get('is_active')).lower() != 'false',
                })
        if not usuarios_validados:
            return Response({'erro': 'Nenhum usuário válido encontrado ou ja cadastrado.', 'status': status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({'usuarios': usuarios_validados, 'status': status.HTTP_200_OK}, status=status.HTTP_200_OK)

class InportadosUsuarios(APIView):
    def post(self, request):
        usuarios = request.data.get('usuarios')
        if not usuarios:
            return Response({'erro': 'Nenhum usuário enviado.', 'status': status.HTTP_400_BAD_REQUEST}, status=status.HTTP_400_BAD_REQUEST)

        cadastrados = 0

        for usuario_data in usuarios:
            try:
                CustomUser.objects.create_user(
                    username=usuario_data['username'],
                    first_name=usuario_data['first_name'],
                    password=make_password(usuario_data['senha']),
                    ra=usuario_data['ra'],
                    saldo=usuario_data['saldo'],
                    pontuacao=usuario_data['pontuacao'],
                    is_adm=usuario_data['is_adm'],
                    is_active=usuario_data['is_active']
                )
                cadastrados += 1
            except Exception as e:
                # Ignora erros individuais e continua com os demais
                continue

        return Response({'message': f'{cadastrados} usuários cadastrados com sucesso!', 'status': status.HTTP_200_OK}, status=status.HTTP_200_OK)
